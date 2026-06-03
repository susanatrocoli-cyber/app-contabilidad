import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter 
from openpyxl.utils.dataframe import dataframe_to_rows  # Asegúrate de importar correctamente esta función
from openpyxl.styles import Font, Alignment
from io import BytesIO
import zipfile


# Interfaz de inicio
def inicio():
    st.title("Selección de configuración")
    option = st.selectbox("Elige una opción:", ["Configurar contactos", "Configurar gastos", "Configurar PC","Configurar Ingresos","Configurar gastos Sage", "Configurar ingresos Sage"])
    return option

#Interfaz 1
def configurar_contactos():
    # Interfaz de usuario con Streamlit
    st.title("Traspaso contactos Quipu a A3")

    # Subida de múltiples archivos
    contacts_files = st.file_uploader("Sube los archivos de A3", type=["xlsx"], key="contactos_file_uploader", accept_multiple_files=True)

    # Input para configurar el número de dígitos de la columna "Código"
    codigo_length = st.number_input("Número de dígitos para la columna 'Código'", min_value=6, max_value=12, value=8)

    # Botón para ejecutar la transformación
    if st.button("Transformar y Descargar"):
        if contacts_files:  # Comprobar si hay archivos subidos
            try:
                output_buffers = []

                # Procesar cada archivo subido
                for contacts_file in contacts_files:
                    original_filename = contacts_file.name
                    base_name, ext = original_filename.rsplit('.', 1)
                    new_filename = f"{base_name}_transformed.{ext}"

                    # Transformar los datos del archivo actual
                    output_files = transfer_data(contacts_file, codigo_length)

                    if output_files:
                        if len(output_files) == 1:
                            # Añadir el archivo transformado a la lista de buffers
                            output_buffers.append((new_filename, output_files[0]))
                        else:
                            # Si hay más de un fragmento, añadir cada uno con su nombre
                            for i, output_file in enumerate(output_files):
                                output_buffers.append((f"{base_name}_transformed_part_{i + 1}.{ext}", output_file))

                if len(output_buffers) == 1:
                    # Si solo se transformó un archivo, descarga directa
                    st.download_button(
                        label="Descargar archivo transformado",
                        data=output_buffers[0][1].getvalue(),
                        file_name=output_buffers[0][0],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    # Si se transformaron múltiples archivos, crear un ZIP con todos ellos
                    buffer = BytesIO()
                    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for filename, file_buffer in output_buffers:
                            zip_file.writestr(filename, file_buffer.getvalue())

                    buffer.seek(0)

                    # Descargar el archivo ZIP
                    st.download_button(
                        label="Descargar todos los archivos transformados",
                        data=buffer,
                        file_name="contacts_transformed.zip",
                        mime="application/zip"
                    )
            except Exception as e:
                st.error(f"Ocurrió un error: {e}")
        else:
            st.error("Por favor, sube al menos un archivo para transformar.")

    # Supongamos que transfer_data es tu función que transforma los datos

def transfer_data(contacts_file, codigo_length):
    try:
        # Simulación de la transformación del archivo para este ejemplo
        df = pd.read_excel(contacts_file)

        # Realizar las transformaciones en los datos aquí (omitido por simplicidad)
        output_files = []

        # Crear un buffer de memoria para almacenar el archivo transformado
        buffer = BytesIO()
        df.to_excel(buffer, index=False)
        buffer.seek(0)
        output_files.append(buffer)

        return output_files

    except Exception as e:
        st.error(f"Se produjo un error: {e}")
        return None


    # Supongamos que transfer_data es tu función que transforma los datos

def transfer_data(contacts_file, codigo_length):
    try:
        # Simulación de la transformación del archivo para este ejemplo
        df = pd.read_excel(contacts_file)

        # Realizar las transformaciones en los datos aquí (omitido por simplicidad)
        output_files = []

        # Crear un buffer de memoria para almacenar el archivo transformado
        buffer = BytesIO()
        df.to_excel(buffer, index=False)
        buffer.seek(0)
        output_files.append(buffer)

        return output_files

    except Exception as e:
        st.error(f"Se produjo un error: {e}")
        return None

def extract_country(nif):
    # Función para extraer el país desde el NIF

    if isinstance(nif, str):
        if nif.startswith('ES'):
            return 'ES'
        elif nif.startswith('FR'):
            return 'FR'
        elif nif.startswith('DE'):
            return 'DE'
        elif nif.startswith('IT'):
            return 'IT'
        elif nif.startswith('PT'):
            return 'PT'
        elif nif.startswith('BE'):
            return 'BE'
        elif nif.startswith('NL'):
            return 'NL'
        elif nif.startswith('LU'):
            return 'LU'
        elif nif.startswith('IE'):
            return 'IE'
        elif nif.startswith('DK'):
            return 'DK'
        elif nif.startswith('SE'):
            return 'SE'
        elif nif.startswith('FI'):
            return 'FI'
        elif nif.startswith('AT'):
            return 'AT'
        elif nif.startswith('GR'):
            return 'GR'
        elif nif.startswith('CY'):
            return 'CY'
        elif nif.startswith('EE'):
            return 'EE'
        elif nif.startswith('LV'):
            return 'LV'
        elif nif.startswith('LT'):
            return 'LT'
        elif nif.startswith('MT'):
            return 'MT'
        elif nif.startswith('PL'):
            return 'PL'
        elif nif.startswith('CZ'):
            return 'CZ'
        elif nif.startswith('SK'):
            return 'SK'
        elif nif.startswith('SI'):
            return 'SI'
        elif nif.startswith('HU'):
            return 'HU'
        elif nif.startswith('RO'):
            return 'RO'
        elif nif.startswith('BG'):
            return 'BG'
        elif nif.startswith('HR'):
            return 'HR'
        elif nif.startswith('GB'):
            return 'GB'
    return 'ES'

def generar_min_max(num_digitos):
    #  Función para generar dígitos contables proveedor mínima y máxima
    if num_digitos < 6 or num_digitos > 12:
        raise ValueError("El número de dígitos debe estar entre 6 y 12.")
    
    # Calculamos el mínimo y el máximo según la cantidad de dígitos
    min_val = int('4' + '0' * (num_digitos - 1))
    
    # Si num_digitos es 10, por ejemplo, max_val será '4999999999'
    max_val = int('4' + '9' * (num_digitos - 1))  # Generamos un número que comienza con 4 y luego '9's
    
    # max_val_cliente será algo como '4300000000' si num_digitos es 10
    max_val_cliente = int('43' + '0' * (num_digitos - 2))
    
    return min_val, max_val, max_val_cliente

def transfer_data(contacts_file, codigo_length):
    try:
        # Cargar el archivo de Reina Campos
        df_reina_campos = pd.read_excel(contacts_file, sheet_name='Cuentas Datos Tesorería')

        # Limpiar los datos de Reina Campos comenzando desde la segunda fila
        df_reina_campos_clean = df_reina_campos.iloc[0:].copy()
        df_reina_campos_clean.columns = ['Código', 'Descripción', 'N.I.F.', 'Contrapartida', 'Unnamed_4', 'Unnamed_5', 'Unnamed_6', 'Unnamed_7', 'Unnamed_8', 'Unnamed_9']
        df_reina_campos_clean = df_reina_campos_clean[['Código', 'Descripción', 'N.I.F.', 'Contrapartida']]

        # Ajustar el formato de la columna "Código" según el número de dígitos especificado
        df_reina_campos_clean['Código'] = df_reina_campos_clean['Código'].astype(str).str.zfill(codigo_length)
        
        df_reina_campos_clean = df_reina_campos_clean.dropna(subset=['Descripción', 'N.I.F.'])

        # Definir las columnas del nuevo DataFrame
        columnas_definidas = ['Nombre', 'Email', 'NIF', 'Dirección', 'Población', 'Código Postal', 'País', 'Teléfono', 'IBAN', 'SWIFT/BIC', 'Cuenta de cliente', 'Cuenta de Proveedor', 'Cuenta de trabajador', 'Contrapartida de ingreso', 'Contrapartida de gasto']
        df_contacts_sample = pd.DataFrame(columns=columnas_definidas)

        # Definir los rangos de códigos y contrapartidas basados en el número de dígitos
        min_contrapartida = 6 ** (codigo_length - 1)
        max_contrapartida = 12 ** (codigo_length - 1)
        min_proveedor, max_proveedor, max_cliente = generar_min_max(codigo_length)

        # Crear una lista para almacenar las filas a añadir
        rows_to_add = []


        # Iterar por cada fila de Reina Campos y preparar los datos para el nuevo DataFrame
        for i, row in df_reina_campos_clean.iterrows():
            new_row = {
                'Nombre': row['Descripción'],
                'Email': '',
                'NIF': row['N.I.F.'],
                'Dirección': '',
                'Población': '',
                'Código Postal': '',
                'País': extract_country(row['N.I.F.']),
                'Teléfono': '',
                'IBAN': '',
                'SWIFT/BIC': '',
                'Cuenta de cliente': '',
                'Cuenta de Proveedor': '',
                'Cuenta de trabajador': '',
                'Contrapartida de ingreso': '',
                'Contrapartida de gasto': '',
            }

            

            # Manejar posibles valores no numéricos en "Código"
            try:
                codigo = int(row['Código'])
                if codigo >= max_cliente:
                    new_row['Cuenta de cliente'] = row['Código']
                    new_row['Contrapartida de ingreso'] = row['Contrapartida']
                elif min_proveedor <= codigo <= max_proveedor:
                    new_row['Cuenta de Proveedor'] = row['Código']
                    new_row['Contrapartida de gasto'] = row['Contrapartida']

            except ValueError:
                pass

            rows_to_add.append(new_row)

        df_contacts_sample = pd.concat([df_contacts_sample, pd.DataFrame(rows_to_add)], ignore_index=True)

        
        # Dividir el DataFrame en fragmentos de hasta 900 filas
        chunk_size = 900
        chunks = [df_contacts_sample[i:i + chunk_size] for i in range(0, df_contacts_sample.shape[0], chunk_size)]

        # Crear una lista de buffers de los archivos generados
        output_files = []
        for i, chunk in enumerate(chunks):
            chunk_buffer = BytesIO()
            chunk.to_excel(chunk_buffer, index=False)
            chunk_buffer.seek(0)
            output_files.append(chunk_buffer)
            
        return output_files

    except Exception as e:
        st.error(f"Se produjo un error: {e}")
        return None
    
#Interfaz 2

# Función para agregar el año 2026 si falta

def aplicar_formato_excel_gastos(df):

    wb = Workbook()
    ws = wb.active

    # Definir encabezados
    headers = [
        'Factura o Ticket',
        'Fecha de emisión',
        'Fecha de vencimiento',
        'Fecha de pago',
        'Estado de pago',
        'Numeración',
        'NIF',
        'Nombre',
        'Correo',
        'Dirección',
        'Población',
        'Código postal',
        'País',
        'Teléfono',
        'Número de cuenta',
        'SWIFT/BIC',
        'Código contable proveedor',
        'Concepto',
        'Base unitaria',
        'Cantidad',
        'IVA (%)',
        'Recargo de equivalencia',
        'Retención (%)',
        'Número de cuenta',
        'Sección',
        'Epígrafe'
    ]

    # Escribir encabezados en la fila 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Combinar celdas en la fila 1 para los encabezados agrupados
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Datos fiscales'
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('G1:Q1')
    ws['G1'] = 'Proveedor'
    ws['G1'].font = Font(bold=True)
    ws['G1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('R1:W1')
    ws['R1'] = 'Línea de factura'
    ws['R1'].font = Font(bold=True)
    ws['R1'].alignment = Alignment(horizontal='center')

    ws['X1'] = 'Categoría de gasto'
    ws['X1'].font = Font(bold=True)
    ws['X1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('Y1:Z1')
    ws['Y1'] = 'Actividad Económica'
    ws['Y1'].font = Font(bold=True)
    ws['Y1'].alignment = Alignment(horizontal='center')

    # Verificación del DataFrame antes de agregarlo al Excel
    print(df.columns)  # Ver las columnas del DataFrame
    print(len(df.columns))  # Ver el número de columnas en el DataFrame

    # Añadir los datos del DataFrame al archivo Excel a partir de la fila 3
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    # Añadir una fila en blanco explícitamente
    ws.append([''] * ws.max_column)

    # Ajustar el área de impresión para que incluya solo el rango necesario
    ws.print_area = f'A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}'

    # Guardar el archivo Excel en un objeto BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)

    # Volver a cargar el archivo para asegurarnos de que las propiedades se recalculen correctamente
    excel_buffer.seek(0)
    wb_reloaded = load_workbook(excel_buffer)
    excel_buffer_reloaded = BytesIO()
    wb_reloaded.save(excel_buffer_reloaded)
    excel_buffer_reloaded.seek(0)

    return excel_buffer_reloaded

def configurar_gastos():
    st.title("Automatización de Traspaso de Excel Gastos")

    # Subida de múltiples archivos
    uploaded_aquacee = st.file_uploader("Sube los archivos históricos de A3", type="xlsx", accept_multiple_files=True, key="gastos_file_uploader")
    
    if uploaded_aquacee:
        st.write("Histórico subido correctamente.")
    
    try:
        buffer = BytesIO()  # Buffer para almacenar el archivo ZIP
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for uploaded_file in uploaded_aquacee:  # Iterar sobre cada archivo cargado
                # Cargar los datos de cada archivo
                aquacee_df = pd.read_excel(uploaded_file)  # Leer el archivo de Excel

                # Procesar los datos
                aquacee_df.columns = ['NºOrden', 'NReferencia', 'NumFact', 'Fecha', 'Concepto', 'NIF', 'Expedidor',
                                      'BaseImponible', '%IVA', 'Cuota', 'Retencion', 'TotalFra']

                # Crear el DataFrame de gastos con los datos transformados
                gastos_df = pd.DataFrame()
                try:
                    gastos_df['Fecha de emisión'] = pd.to_datetime(aquacee_df['Fecha'].astype(str) + "/2026", format="%d/%m/%Y", errors='coerce').dt.strftime('%d-%m-%Y')
                except Exception as fecha_error:
                    st.error(f"Error al convertir las fechas: {fecha_error}")
                    gastos_df['Fecha de emisión'] = ''
                gastos_df['NIF'] = aquacee_df['NIF']
                gastos_df['Numeración'] = aquacee_df['NumFact']
                gastos_df['Nombre'] = aquacee_df['Expedidor']
                gastos_df['Concepto'] = "Gastos de " + aquacee_df['Expedidor'].astype(str)
                gastos_df['Base unitaria'] = aquacee_df['BaseImponible']
                gastos_df['Cantidad'] = 1
                gastos_df['IVA (%)'] = aquacee_df['%IVA']
                gastos_df['Retención (%)'] = aquacee_df.apply(
                    lambda row: round(row['Retencion'] / row['BaseImponible'] * 100, 2) if row['Retencion'] != 0 else 0, axis=1
                )
                gastos_df['Estado de pago'] = 'Pendiente'
                gastos_df['Factura o Ticket'] = aquacee_df.apply(
                    lambda row: 'Factura' if pd.notna(row['NIF']) else 'Ticket', axis=1
                )
                gastos_df['País'] = gastos_df['NIF'].apply(lambda x: extract_country(x) if isinstance(x, str) and len(x) > 0 else 'ES')

                # Asegurarse de que todas las columnas estén en el DataFrame, incluso si no tienen datos
                column_order = [
                    'Factura o Ticket', 'Fecha de emisión', 'Fecha de vencimiento', 'Fecha de pago', 'Estado de pago',
                    'Numeración', 'NIF', 'Nombre', 'Correo', 'Dirección', 'Población', 'Código postal', 'País',
                    'Teléfono', 'Número de cuenta', 'SWIFT/BIC', 'Código contable proveedor', 'Concepto',
                    'Base unitaria', 'Cantidad', 'IVA (%)', 'Recargo de equivalencia', 'Retención (%)', 'Número de cuenta',
                    'Sección', 'Epígrafe'
                ]
                for col in column_order:
                    if col not in gastos_df.columns:
                        gastos_df[col] = ''

                gastos_df = gastos_df[column_order]

                # Dividir en chunks si es necesario (más de 900 filas por archivo)
                chunk_size = 900
                chunks = [gastos_df[i:i + chunk_size] for i in range(0, gastos_df.shape[0], chunk_size)]

                for i, chunk in enumerate(chunks):
                    chunk_buffer = aplicar_formato_excel_gastos(chunk)

                    # Crear un nombre para el archivo basado en el nombre original del archivo subido
                    original_filename = uploaded_file.name.split(".")[0]
                    chunk_filename = f"{original_filename}_transformed_parte_{i + 1}.xlsx"

                    zip_file.writestr(chunk_filename, chunk_buffer.getvalue())

        buffer.seek(0)

        # Botón para descargar el archivo ZIP
        st.download_button(
            label="Descargar archivos ZIP con Excel divididos",
            data=buffer,
            file_name="gastos_divididos.zip",
            mime="application/zip"
        )

    except Exception as e:
        st.error(f"Se produjo un error durante el procesamiento: {e}")

#Interfaz 3
def generar_min_max_2(num_digitos):
    if num_digitos < 6 or num_digitos > 12:
        raise ValueError("El número de dígitos debe estar entre 6 y 12.")
    
    # Generar rangos para los códigos en función del número de dígitos
    min_val_20_to_29 = int('2' + '0' * (num_digitos - 1))
    max_val_20_to_29 = int('2' + '9' * (num_digitos - 1))
    
    min_val_60_to_79 = int('6' + '0' * (num_digitos - 1))
    max_val_60_to_79 = int('7' + '9' * (num_digitos - 1))
    
    return min_val_20_to_29, max_val_20_to_29, min_val_60_to_79, max_val_60_to_79

def automate_excel_transfer(df_entalud, num_digitos):
    # Obtener los rangos de los códigos según la cantidad de dígitos
    min_20_to_29, max_20_to_29, min_60_to_79, max_60_to_79 = generar_min_max_2(num_digitos)

    # Asegurarse de que la columna de códigos sea numérica
    df_entalud.iloc[:, 0] = pd.to_numeric(df_entalud.iloc[:, 0], errors='coerce')

    # Filtrar los códigos en función del número de dígitos y los rangos
    codes_20_to_29 = df_entalud[(df_entalud.iloc[:, 0] >= min_20_to_29) & (df_entalud.iloc[:, 0] <= max_20_to_29)]
    codes_60_to_79 = df_entalud[(df_entalud.iloc[:, 0] >= min_60_to_79) & (df_entalud.iloc[:, 0] <= max_60_to_79)]

    # Combinar los datos filtrados
    combined_codes = pd.concat([codes_20_to_29, codes_60_to_79])

    # Crear un nuevo DataFrame con los encabezados 'Cuenta' y 'Descripción'
    filtered_data = combined_codes.iloc[:, [0, 1]].copy()
    filtered_data.columns = ['Cuenta', 'Descripción']

    return filtered_data

def Configurar_PC():
    
  # Interfaz de Streamlit
    st.title("Automatización de Transferencia de Planes Contables")

    # Subir múltiples archivos de Entalud
    entalud_files = st.file_uploader("Sube los planes contables de A3", type=["xlsx"], accept_multiple_files=True)

    # Número de dígitos para los códigos contables
    num_digitos = st.number_input("Número de dígitos para los códigos contables", min_value=6, max_value=12, value=8)

    # Botón para ejecutar la transferencia y generar ZIP
    if st.button("Generar ZIP con archivos transformados"):
        if entalud_files:
            zip_buffer = BytesIO()

            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                for entalud_file in entalud_files:
                    # Leer el archivo subido
                    df_entalud = pd.read_excel(entalud_file, header=None)
                    filtered_data = automate_excel_transfer(df_entalud, num_digitos)
                    
                    # Generar un archivo Excel en memoria
                    original_name = entalud_file.name.split('.')[0]
                    transformed_filename = f"{original_name}-transformed.xlsx"
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        filtered_data.to_excel(writer, index=False)
                    output.seek(0)
                    
                    # Añadir el archivo Excel al ZIP
                    zipf.writestr(transformed_filename, output.getvalue())

            # Descargar el archivo ZIP
            zip_buffer.seek(0)
            st.download_button(
                label="Descargar ZIP",
                data=zip_buffer,
                file_name="transformed_files.zip",
                mime="application/zip"
            )
        else:
            st.warning("Por favor, sube al menos un archivo antes de generar el ZIP.")

#Interfaz 4:
def aplicar_formato_excel_ingresos(df):
    # Función para aplicar formato a cada archivo de ingresos
    wb = Workbook()
    ws = wb.active

    # Definir encabezados
    headers = [
        'Factura, Ticket',
        'Fecha de emisión',
        'Fecha de vencimiento',
        'Fecha de pago',
        'Estado de pago',
        'Serie de facturación',
        'Número de ticket o factura',
        'NIF',
        'Nombre',
        'Correo',
        'Dirección',
        'Población',
        'Código postal',
        'País',
        'Teléfono',
        'IBAN',
        'SWIFT/BIC',
        'Código contable cliente',
        'Concepto',
        'Base unitaria',
        'Cantidad',
        'IVA (%)',
        'Recargo de equivalencia',
        'Retención (%)',
        'Código contable',
        'Sección',
        'Epígrafe',
        'Tipo'
    ]

    # Escribir encabezados en la fila 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Combinar celdas en la fila 1 para los encabezados agrupados
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Datos fiscales'
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('H1:R1')
    ws['H1'] = 'Cliente'
    ws['H1'].font = Font(bold=True)
    ws['H1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('S1:X1')
    ws['S1'] = 'Línea de factura'
    ws['S1'].font = Font(bold=True)
    ws['S1'].alignment = Alignment(horizontal='center')

    ws['Y1'] = 'Categoría de ingreso'
    ws['Y1'].font = Font(bold=True)
    ws['Y1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('Z1:AA1')
    ws['Z1'] = 'Actividad Económica'
    ws['Z1'].font = Font(bold=True)
    ws['Z1'].alignment = Alignment(horizontal='center')

    ws['AB1'] = 'Tipo de cliente'
    ws['AB1'].font = Font(bold=True)
    ws['AB1'].alignment = Alignment(horizontal='center')

    # Añadir los datos del DataFrame al archivo Excel a partir de la fila 3
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Añadir una fila en blanco explícitamente
    ws.append([''] * ws.max_column)

    # Ajustar el área de impresión para que incluya solo el rango necesario
    ws.print_area = f'A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}'

    # Guardar el archivo Excel en un objeto BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)

    # Volver a cargar el archivo para asegurarnos de que las propiedades se recalculen correctamente
    excel_buffer.seek(0)
    wb_reloaded = load_workbook(excel_buffer)
    excel_buffer_reloaded = BytesIO()
    wb_reloaded.save(excel_buffer_reloaded)
    excel_buffer_reloaded.seek(0)

    return excel_buffer_reloaded

def configurar_ingresos():
    st.title("Automatización de Traspaso de Excel Ingresos")

    # Subida de archivo AQUACEE RECIBIDAS (múltiples archivos)
    uploaded_aquacee = st.file_uploader("Sube los archivos históricos de A3", type="xlsx", accept_multiple_files=True, key="ingresos_file_uploader")

    if uploaded_aquacee:
        try:
            buffer = BytesIO()  # Buffer para almacenar el archivo ZIP
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for uploaded_file in uploaded_aquacee:  # Iterar sobre cada archivo cargado
                    # Cargar los datos de cada archivo
                    aquacee_df = pd.read_excel(uploaded_file, skiprows=0)

                    # Saltar encabezados innecesarios
                    aquacee_df.columns = ['NºOrden', 'NReferencia', 'NumFact', 'Fecha', 'Concepto', 'NIF', 'Destinatario',
                                          'BaseImponible', '%IVA', 'Cuota', 'Retencion', 'TotalFra']

                    # Crear el DataFrame de ingresos con los datos transformados
                    ingresos_df = pd.DataFrame()
                    # Manejar posibles errores en la conversión de fechas
                    try:
                        ingresos_df['Fecha de emisión'] = pd.to_datetime(aquacee_df['Fecha'].astype(str) + "/2026", format="%d/%m/%Y", errors='coerce').dt.strftime('%d-%m-%Y')
                    except Exception as fecha_error:
                        st.error(f"Error al convertir las fechas: {fecha_error}")
                        ingresos_df['Fecha de emisión'] = ''
                    ingresos_df['Número de ticket o factura'] = aquacee_df['NumFact']
                    ingresos_df['NIF'] = aquacee_df['NIF']
                    ingresos_df['Nombre'] = aquacee_df['Destinatario']
                    ingresos_df['Concepto'] = "Ingresos de " + aquacee_df['Destinatario'].astype(str)
                    ingresos_df['Base unitaria'] = aquacee_df['BaseImponible']
                    ingresos_df['Cantidad'] = 1
                    ingresos_df['IVA (%)'] = aquacee_df['%IVA']
                    ingresos_df['Retención (%)'] = aquacee_df.apply(
                        lambda row: round(row['Retencion'] / row['BaseImponible'] * 100, 2) if row['Retencion'] != 0 else "", axis=1)
                    ingresos_df['Estado de pago'] = 'Pendiente'
                    ingresos_df['Factura, Ticket'] = aquacee_df.apply(
                        lambda row: 'Factura' if pd.notna(row['NIF']) else 'Ticket', axis=1
                    )
                    ingresos_df['País'] = aquacee_df['NIF'].apply(lambda x: extract_country(x) if isinstance(x, str) and len(x) > 0 else 'ES')

                    # Asegúrate de que todas las columnas estén en el DataFrame, incluso si no tienen datos
                    column_order = [
                        'Factura, Ticket', 'Fecha de emisión', 'Fecha de vencimiento', 'Fecha de pago', 'Estado de pago',
                        'Serie de facturación', 'Número de ticket o factura', 'NIF', 'Nombre', 'Correo', 'Dirección',
                        'Población', 'Código postal', 'País', 'Teléfono', 'IBAN', 'SWIFT/BIC', 'Código contable cliente',
                        'Concepto', 'Base unitaria', 'Cantidad', 'IVA (%)', 'Recargo de equivalencia', 'Retención (%)',
                        'Código contable', 'Sección', 'Epígrafe', 'Tipo'
                    ]
                    for col in column_order:
                        if col not in ingresos_df.columns:
                            ingresos_df[col] = ''

                    ingresos_df = ingresos_df[column_order]

                    st.write(f"Datos procesados para {uploaded_file.name} y preparados para exportar.")

                    # Crear un archivo ZIP con los datos procesados
                    chunk_size = 900  # Tamaño de cada parte
                    chunks = [ingresos_df[i:i + chunk_size] for i in range(0, ingresos_df.shape[0], chunk_size)]

                    for i, chunk in enumerate(chunks):
                        # Aplicar formato a cada archivo de chunk
                        chunk_buffer = aplicar_formato_excel_ingresos(chunk)
                        zip_file.writestr(f'{uploaded_file.name.split(".")[0]}_parte_{i + 1}.xlsx', chunk_buffer.getvalue())

            buffer.seek(0)

            # Botón para descargar el archivo ZIP
            st.download_button(
                label="Descargar archivos ZIP con Excel divididos",
                data=buffer,
                file_name="ingresos_divididos.zip",
                mime="application/zip"
            )
        except Exception as e:
            st.error(f"Se produjo un error durante el procesamiento: {e}")

    else:
        st.write("Por favor, sube el histórico de A3 para continuar.")

#Interfaz SAGE- Gastos:

def aplicar_formato_excel_gastos_sage(df):
    wb = Workbook()
    ws = wb.active

    # Definir encabezados
    headers = [
        'Factura o Ticket',
        'Fecha de emisión',
        'Fecha de vencimiento',
        'Fecha de pago',
        'Estado de pago',
        'Numeración',
        'NIF',
        'Nombre',
        'Correo',
        'Dirección',
        'Población',
        'Código postal',
        'País',
        'Teléfono',
        'Número de cuenta',
        'SWIFT/BIC',
        'Código contable proveedor',
        'Concepto',
        'Base unitaria',
        'Cantidad',
        'IVA (%)',
        'Recargo de equivalencia',
        'Retención (%)',
        'Número de cuenta',
        'Sección',
        'Epígrafe'
    ]

    # Escribir encabezados en la fila 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Combinar celdas en la fila 1 para los encabezados agrupados
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Datos fiscales'
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('G1:Q1')
    ws['G1'] = 'Proveedor'
    ws['G1'].font = Font(bold=True)
    ws['G1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('R1:W1')
    ws['R1'] = 'Línea de factura'
    ws['R1'].font = Font(bold=True)
    ws['R1'].alignment = Alignment(horizontal='center')

    ws['X1'] = 'Categoría de gasto'
    ws['X1'].font = Font(bold=True)
    ws['X1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('Y1:Z1')
    ws['Y1'] = 'Actividad Económica'
    ws['Y1'].font = Font(bold=True)
    ws['Y1'].alignment = Alignment(horizontal='center')

    # Añadir los datos del DataFrame al archivo Excel a partir de la fila 3
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Añadir una fila en blanco explícitamente
    ws.append([''] * ws.max_column)

    # Ajustar el área de impresión para que incluya solo el rango necesario
    ws.print_area = f'A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}'

    # Guardar el archivo Excel en un objeto BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)

    # Volver a cargar el archivo para asegurarnos de que las propiedades se recalculen correctamente
    excel_buffer.seek(0)
    wb_reloaded = load_workbook(excel_buffer)
    excel_buffer_reloaded = BytesIO()
    wb_reloaded.save(excel_buffer_reloaded)
    excel_buffer_reloaded.seek(0)

    return excel_buffer_reloaded

def configurar_gastos_sage():
    st.title("Automatización de Traspaso de Excel Gasto Sage")
    
    # Subida de múltiples archivos
    uploaded_gastos_sage = st.file_uploader("Sube los archivos históricos de A3", type="xlsx", accept_multiple_files=True, key="gastos_file_uploader")
    
    if uploaded_gastos_sage:
        st.write("Histórico subido correctamente.")
        try:
            buffer = BytesIO()  # Buffer para almacenar el archivo ZIP
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for uploaded_file in uploaded_gastos_sage:  # Iterar sobre cada archivo cargado
                    # Cargar los datos de cada archivo desde la segunda fila
                    sage_df = pd.read_excel(uploaded_file, header=1)  # Leer el archivo de Excel desde la segunda fila
                    
                    # Depuración: imprimir las columnas reales del archivo cargado
                    st.write(f"El archivo {uploaded_file.name} tiene {sage_df.shape[1]} columnas.")
                    st.write(f"Columnas del archivo: {sage_df.columns.tolist()}")

                    # Lista esperada de columnas (23 columnas en tu caso)
                    expected_columns = ['Ejercicio', 'Periodo', 'Tipo', 'Grupo o Epígrafe del IAE', 'Tipo de Factura', 'Concepto de Gasto', 'Gasto Deducible', 'Fecha Expedición', 
                                        'Fecha Operación', '(Serie-Número)', 'Número-Final', 'Número Recepción', 'Número Recepción Final', 'Tipo.1', 'Código País', 'Identificación', 'Nombre Expedidor', 
                                        'Clave de Operación', 'Total Factura', 'Base Imponible', 'Tipo de IVA', 'Cuota IVA Soportado', 'Cuota Deducible', 'Tipo de Recargo Eq.', 'Cuota Recargo Eq.',
                                         'Fecha', 'Importe', 'Medio Utilizado', 'Identificación Medio Utilizado', 'Tipo Retención del IRPF', 'Importe Retenido del IRPF']

                    # Verificar si las columnas coinciden
                    if len(sage_df.columns) != len(expected_columns):
                        st.error(f"El archivo {uploaded_file.name} tiene un número de columnas diferente al esperado. Esperado: {len(expected_columns)}, Recibido: {len(sage_df.columns)}")
                        continue

                    # Asignar los nombres esperados si el número de columnas coincide
                    sage_df.columns = expected_columns
                    
                    # Crear el DataFrame de gastos_sage con los datos transformados
                    gastos_sage_df = pd.DataFrame()
                    gastos_sage_df['Fecha de emisión'] = pd.to_datetime(sage_df['Fecha Expedición'], errors='coerce').dt.strftime('%d-%m-%Y').fillna('')
                    gastos_sage_df['Nombre'] = sage_df['Nombre Expedidor']
                    gastos_sage_df['NIF'] = sage_df['Identificación']
                    gastos_sage_df['Numeración'] = sage_df['(Serie-Número)']
                    gastos_sage_df['Base unitaria'] = sage_df['Base Imponible']
                    gastos_sage_df['Cantidad'] = 1
                    gastos_sage_df['IVA (%)'] = sage_df['Tipo de IVA'].fillna(0)
                    gastos_sage_df['Retención (%)'] = sage_df['Tipo Retención del IRPF'].fillna(0)
                    gastos_sage_df['Estado de pago'] = 'Pendiente'
                    gastos_sage_df['Factura o Ticket'] = gastos_sage_df.apply(
                    lambda row: 'Factura' if isinstance(row['NIF'], str) and len(row['NIF']) > 0 else 'Ticket', axis=1)

                    gastos_sage_df['País'] = gastos_sage_df['NIF'].apply(lambda x: extract_country(x) if isinstance(x, str) and len(x) > 0 else 'ES')
                    # Asegurarse de que todas las columnas estén en el DataFrame
                    column_order = [
                        'Factura o Ticket', 'Fecha de emisión', 'Fecha de vencimiento', 'Fecha de pago', 'Estado de pago',
                        'Numeración', 'NIF', 'Nombre', 'Correo', 'Dirección', 'Población', 'Código postal', 'País',
                        'Teléfono', 'Número de cuenta', 'SWIFT/BIC', 'Código contable proveedor', 'Concepto',
                        'Base unitaria', 'Cantidad', 'IVA (%)', 'Recargo de equivalencia', 'Retención (%)', 'Número de cuenta',
                        'Sección', 'Epígrafe'
                    ]

                    for col in column_order:
                        if col not in gastos_sage_df.columns:
                            gastos_sage_df[col] = ''

                    gastos_sage_df = gastos_sage_df[column_order]

                    # Dividir en chunks si es necesario (más de 900 filas por archivo)
                    chunk_size = 900
                    chunks = [gastos_sage_df[i:i + chunk_size] for i in range(0, gastos_sage_df.shape[0], chunk_size)]

                    for i, chunk in enumerate(chunks):
                        chunk_buffer = aplicar_formato_excel_gastos_sage(chunk)

                        # Crear un nombre para el archivo basado en el nombre original del archivo subido
                        original_filename = uploaded_file.name.split(".")[0]
                        chunk_filename = f"{original_filename}_transformed_parte_{i + 1}.xlsx"

                        zip_file.writestr(chunk_filename, chunk_buffer.getvalue())

            buffer.seek(0)

            # Botón para descargar el archivo ZIP
            st.download_button(
                label="Descargar archivos ZIP con Excel divididos",
                data=buffer,
                file_name="gastos_divididos.zip",
                mime="application/zip"
            )

        except Exception as e:
            st.error(f"Se produjo un error durante el procesamiento: {e}")

#Interfaz SAGE- Ingresos:

def aplicar_formato_excel_ingresos_sage(df):
    # Función para aplicar formato a cada archivo de ingresos
    wb = Workbook()
    ws = wb.active

    # Definir encabezados
    headers = [
        'Factura, Ticket',
        'Fecha de emisión',
        'Fecha de vencimiento',
        'Fecha de pago',
        'Estado de pago',
        'Serie de facturación',
        'Número de ticket o factura',
        'NIF',
        'Nombre',
        'Correo',
        'Dirección',
        'Población',
        'Código postal',
        'País',
        'Teléfono',
        'IBAN',
        'SWIFT/BIC',
        'Código contable cliente',
        'Concepto',
        'Base unitaria',
        'Cantidad',
        'IVA (%)',
        'Recargo de equivalencia',
        'Retención (%)',
        'Código contable',
        'Sección',
        'Epígrafe',
        'Tipo'
    ]

    # Escribir encabezados en la fila 2
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Combinar celdas en la fila 1 para los encabezados agrupados
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Datos fiscales'
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('H1:R1')
    ws['H1'] = 'Cliente'
    ws['H1'].font = Font(bold=True)
    ws['H1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('S1:X1')
    ws['S1'] = 'Línea de factura'
    ws['S1'].font = Font(bold=True)
    ws['S1'].alignment = Alignment(horizontal='center')

    ws['Y1'] = 'Categoría de ingreso'
    ws['Y1'].font = Font(bold=True)
    ws['Y1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('Z1:AA1')
    ws['Z1'] = 'Actividad Económica'
    ws['Z1'].font = Font(bold=True)
    ws['Z1'].alignment = Alignment(horizontal='center')

    ws['AB1'] = 'Tipo de cliente'
    ws['AB1'].font = Font(bold=True)
    ws['AB1'].alignment = Alignment(horizontal='center')

def configurar_ingresos_sage():
    st.title("Automatización de Traspaso de Excel Gasto Sage")
    
    # Subida de múltiples archivos
    uploaded_ingresos_sage = st.file_uploader("Sube los archivos históricos de A3", type="xlsx", accept_multiple_files=True, key="gastos_file_uploader")
    
    if uploaded_ingresos_sage:
        st.write("Histórico subido correctamente.")
        try:
            buffer = BytesIO()  # Buffer para almacenar el archivo ZIP
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for uploaded_file in uploaded_ingresos_sage:  # Iterar sobre cada archivo cargado
                    # Cargar los datos de cada archivo desde la segunda fila
                    sage_df = pd.read_excel(uploaded_file, header=1)  # Leer el archivo de Excel desde la segunda fila
                    
                    # Depuración: imprimir las columnas reales del archivo cargado
                    st.write(f"El archivo {uploaded_file.name} tiene {sage_df.shape[1]} columnas.")
                    st.write(f"Columnas del archivo: {sage_df.columns.tolist()}")

                    # Lista esperada de columnas (23 columnas en tu caso)
                    expected_columns = ['Ejercicio', 'Periodo', 'Tipo', 'Grupo o Epígrafe del IAE', 'Tipo de Factura', 'Concepto de Ingreso', 'Ingreso Computable', 
                                        'Fecha Expedición', 'Fecha Operación', 'Serie', 'Número', 'Número-Final', 'Tipo.1', 'Código País', 'Identificación', 'Nombre Destinatario',
                                        'Clave de Operación', 'Total Factura', 'Base Imponible', 'Tipo de IVA', 'Cuota IVA Repercutida', 'Tipo de Recargo Eq.', 'Cuota Recargo Eq.',
                                        'Fecha', 'Importe', 'Medio Utilizado', 'Identificación Medio Utilizado', 'Tipo Retención del IRPF', 'Importe Retenido del IRPF']

                    # Verificar si las columnas coinciden
                    if len(sage_df.columns) != len(expected_columns):
                        st.error(f"El archivo {uploaded_file.name} tiene un número de columnas diferente al esperado. Esperado: {len(expected_columns)}, Recibido: {len(sage_df.columns)}")
                        continue

                    # Asignar los nombres esperados si el número de columnas coincide
                    sage_df.columns = expected_columns
                    
                    # Crear el DataFrame de gastos_sage con los datos transformados
                    ingresos_sage_df = pd.DataFrame()
                    ingresos_sage_df['Fecha de emisión'] = pd.to_datetime(sage_df['Fecha Expedición'], errors='coerce').dt.strftime('%d-%m-%Y').fillna('')
                    ingresos_sage_df['Nombre'] = sage_df['Nombre Destinatario']
                    ingresos_sage_df['NIF'] = sage_df['Identificación']
                    ingresos_sage_df['Numeración'] = sage_df['Número']
                    ingresos_sage_df['Base unitaria'] = sage_df['Base Imponible']
                    ingresos_sage_df['Cantidad'] = 1
                    ingresos_sage_df['IVA (%)'] = sage_df['Tipo de IVA'].fillna(0)
                    ingresos_sage_df['Retención (%)'] = sage_df['Tipo Retención del IRPF'].fillna(0)
                    ingresos_sage_df['Estado de pago'] = 'Pendiente'
                    ingresos_sage_df['Factura o Ticket'] = ingresos_sage_df.apply(
                    lambda row: 'Factura' if isinstance(row['NIF'], str) and len(row['NIF']) > 0 else 'Ticket', axis=1)

                    ingresos_sage_df['País'] = ingresos_sage_df['NIF'].apply(lambda x: extract_country(x) if isinstance(x, str) and len(x) > 0 else 'ES')
                    # Asegurarse de que todas las columnas estén en el DataFrame
                    column_order = [
                        'Factura o Ticket', 'Fecha de emisión', 'Fecha de vencimiento', 'Fecha de pago', 'Estado de pago',
                        'Numeración', 'NIF', 'Nombre', 'Correo', 'Dirección', 'Población', 'Código postal', 'País',
                        'Teléfono', 'Número de cuenta', 'SWIFT/BIC', 'Código contable proveedor', 'Concepto',
                        'Base unitaria', 'Cantidad', 'IVA (%)', 'Recargo de equivalencia', 'Retención (%)', 'Número de cuenta',
                        'Sección', 'Epígrafe','Tipo'
                    ]

                    for col in column_order:
                        if col not in ingresos_sage_df.columns:
                            ingresos_sage_df[col] = ''

                    ingresos_sage_df = ingresos_sage_df[column_order]

                    # Dividir en chunks si es necesario (más de 900 filas por archivo)
                    chunk_size = 900
                    chunks = [ingresos_sage_df[i:i + chunk_size] for i in range(0, ingresos_sage_df.shape[0], chunk_size)]

                    for i, chunk in enumerate(chunks):
                        chunk_buffer = aplicar_formato_excel_gastos_sage(chunk)

                        # Crear un nombre para el archivo basado en el nombre original del archivo subido
                        original_filename = uploaded_file.name.split(".")[0]
                        chunk_filename = f"{original_filename}_transformed_parte_{i + 1}.xlsx"

                        zip_file.writestr(chunk_filename, chunk_buffer.getvalue())

            buffer.seek(0)

            # Botón para descargar el archivo ZIP
            st.download_button(
                label="Descargar archivos ZIP con Excel divididos",
                data=buffer,
                file_name="gastos_divididos.zip",
                mime="application/zip"
            )

        except Exception as e:
            st.error(f"Se produjo un error durante el procesamiento: {e}")


# Función principal que controla qué pantalla mostrar
def main():
    option = inicio()
    
    if option == "Configurar contactos":
        configurar_contactos()
    elif option == "Configurar gastos":
        configurar_gastos()
    elif option == "Configurar PC": 
        Configurar_PC()
    elif option == "Configurar Ingresos": 
        configurar_ingresos()
    elif option == "Configurar gastos Sage":
        configurar_gastos_sage()
    elif option == "Configurar ingresos Sage":
        configurar_ingresos_sage()

if __name__ == "__main__":
    main()
